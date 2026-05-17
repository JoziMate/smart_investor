import openpyxl
from openpyxl.utils import column_index_from_string
import logging
from datetime import datetime
import os
import pandas as pd
from config_manager import config
import logger

logger_inst = logging.getLogger(__name__)

ASSET_MAPPING = config["ASSET_MAPPING"]

class ExcelHandler:
    """
    A class to safely handle opening, updating, and saving an Excel workbook.
    """

    def __init__(self, filename: str, sheet_name: str):
        """
        Initializes the ExcelHandler.

        Args:
            filename (str): The path to the Excel file.
            sheet_name (str): The name of the worksheet to modify.
        """
        self.filename = filename
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None

    def load_workbook(self) -> bool:
        """
        Loads the workbook and selects the target sheet.

        Returns:
            bool: True if successful, False otherwise.
        """
        try:
            # Load the workbook, keeping formulas and formatting intact
            self.workbook = openpyxl.load_workbook(self.filename)

            if self.sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[self.sheet_name]
                logger_inst.info(
                    f"Successfully loaded sheet '{
                        self.sheet_name}' from {
                        self.filename}")
                return True
            else:
                logger_inst.error(
                    f"Sheet '{
                        self.sheet_name}' not found in {
                        self.filename}. Available sheets: {
                        self.workbook.sheetnames}")
                return False

        except FileNotFoundError:
            logger_inst.error(f"File not found: {self.filename}")
            return False
        except PermissionError:
            logger_inst.error(
                f"Permission denied: {
                    self.filename}. Is the file open in another program?")
            return False
        except Exception as e:
            logger_inst.error(f"Unexpected error loading {self.filename}: {e}")
            return False

    def update_asset(
            self,
            asset: str,
            current_price: float) -> bool:
        """
        Updates the date and price cells for a specific asset based on ASSET_MAPPING.

        Args:
            asset (str): The asset ticker (e.g., 'MSFT').
            current_price (float): The fetched current market price.

        Returns:
            bool: True if the update was successful, False otherwise.
        """
        if self.sheet is None:
            logger_inst.error(
                "Cannot update asset: Workbook or sheet is not loaded.")
            return False

        if asset not in ASSET_MAPPING:
            logger_inst.warning(
                f"Asset '{asset}' not found in ASSET_MAPPING configuration.")
            return False

        mapping = ASSET_MAPPING[asset]
        row = mapping['row']

        try:
            # Convert column letters to 1-based indexes for openpyxl
            date_col_idx = column_index_from_string(mapping['date_col'])
            price_col_idx = column_index_from_string(mapping['price_col'])

            # Get current date in YYYY-MM-DD format
            today_str = datetime.now().strftime('%Y-%m-%d')

            # Update the cells
            self.sheet.cell(row=row, column=date_col_idx).value = today_str
            self.sheet.cell(
                row=row, column=price_col_idx).value = current_price
            logger_inst.info(
                f"Updated {asset} in Excel: Date={today_str}, Price=${
                    current_price:.2f}")
            return True

        except Exception as e:
            logger_inst.error(f"Failed to update cells for {asset}: {e}")
            return False

    def append_new_position(self, trade_data: dict) -> bool:
        """
        Appends new trade data to the first completely empty row (based on empty Column A)
        in the current sheet.

        Args:
            trade_data (dict): The trade data with keys 'Platform', 'Ticker', 'OrderType', 'Direction', 'Volume', 'Price', 'Justification'.

        Returns:
            bool: True if successful, False otherwise.
        """
        if self.sheet is None:
            logger_inst.error(
                "Cannot append position: Workbook or sheet is not loaded.")
            return False

        try:
            # Date format
            today_str = datetime.now().strftime('%Y-%m-%d')
            ticker_to_add = trade_data.get('Ticker', '')

            # Check for duplicates and find the first empty row
            # Start from row 5 as per "Trejdy" sheet requirements.
            current_row = 5
            while self.sheet.cell(row=current_row, column=1).value is not None:
                existing_date = str(self.sheet.cell(row=current_row, column=1).value).strip()
                existing_ticker = str(self.sheet.cell(row=current_row, column=3).value).strip()

                # Check for duplicate Ticker and Date
                # Existing date might be formatted differently or contain time, so we check if today_str is in existing_date
                if existing_ticker == ticker_to_add and today_str in existing_date:
                    logger_inst.warning(f"Duplicate trade detected for Ticker '{ticker_to_add}' on Date '{today_str}'. Aborting append.")
                    return False

                current_row += 1

            empty_row = current_row

            # Map columns according to requirements:
            # Column A = Date
            # Column B = Platform
            # Column C = Ticker
            # Column D = Order Type
            # Column E = Direction
            # Column F = Volume
            # Column G = Price
            # Column I = Justification

            self.sheet.cell(row=empty_row, column=1).value = today_str
            self.sheet.cell(row=empty_row, column=2).value = trade_data.get('Platform', '')
            self.sheet.cell(row=empty_row, column=3).value = trade_data.get('Ticker', '')
            self.sheet.cell(row=empty_row, column=4).value = trade_data.get('OrderType', '')
            self.sheet.cell(row=empty_row, column=5).value = trade_data.get('Direction', '')
            self.sheet.cell(row=empty_row, column=6).value = trade_data.get('Volume', '')
            self.sheet.cell(row=empty_row, column=7).value = trade_data.get('Price', '')
            self.sheet.cell(row=empty_row, column=9).value = trade_data.get('Justification', '')

            logger_inst.info(
                f"Appended new position at row {empty_row}: {trade_data}")
            return True

        except Exception as e:
            logger_inst.error(f"Failed to append new position: {e}")
            return False

    def update_saldo(self, df_open_pos: pd.DataFrame, usd_pln_rate: float = 1.0) -> bool:
        """
        Calculates balances from 'Pozycje otwarte' and appends them to 'Salda'.
        Multiplies USD-based platforms by usd_pln_rate to convert to PLN.

        Args:
            df_open_pos (pd.DataFrame): DataFrame of the 'Pozycje otwarte' sheet.
            usd_pln_rate (float): The current USD/PLN exchange rate.

        Returns:
            bool: True if successful, False otherwise.
        """
        if self.sheet is None:
            logger_inst.error("Cannot update saldo: Workbook or sheet is not loaded.")
            return False

        try:
            # Initialize balances
            balances = {
                "GPW Trader": 0.0,
                "Saxo": 0.0,
                "Binance Futures": 0.0,
                "Interactive Brokers": 0.0
            }

            # Calculate balances
            for _, row in df_open_pos.iterrows():
                try:
                    platform = str(row.get('Platforma', '')).strip()
                    volume_str = str(row.get('Wolumen', '0')).replace(',', '.')
                    price_str = str(row.get('Cena bieżąca', '0')).replace(',', '.')

                    if not platform or not volume_str or not price_str:
                        continue

                    volume = float(volume_str)
                    price = float(price_str)
                    value = volume * price

                    # Match platform names
                    for key in balances.keys():
                        if key.lower() in platform.lower():
                            balances[key] += value
                            break
                except ValueError:
                    # Skip rows where volume or price cannot be parsed as floats
                    continue

            # Apply USD/PLN conversion for USD platforms
            if "Interactive Brokers" in balances:
                balances["Interactive Brokers"] *= usd_pln_rate
            if "Saxo" in balances:
                balances["Saxo"] *= usd_pln_rate
            if "Binance Futures" in balances:
                balances["Binance Futures"] *= usd_pln_rate

            # Find the first empty row in 'Salda' (starting from row 8 based on skipping 6 headers, 1st data row is 8)
            empty_row = 8
            while self.sheet.cell(row=empty_row, column=1).value is not None:
                empty_row += 1

            today_str = datetime.now().strftime('%Y-%m-%d')

            # Populate the new row
            self.sheet.cell(row=empty_row, column=1).value = "Automatyczny Update"
            self.sheet.cell(row=empty_row, column=2).value = today_str
            self.sheet.cell(row=empty_row, column=3).value = balances["GPW Trader"]
            self.sheet.cell(row=empty_row, column=5).value = balances["Saxo"]
            self.sheet.cell(row=empty_row, column=9).value = balances["Binance Futures"]
            self.sheet.cell(row=empty_row, column=11).value = balances["Interactive Brokers"]

            logger_inst.info(f"Appended new saldo update at row {empty_row}.")

            # Update Historia sheet with the new total portfolio value
            total_pln = sum(balances.values())
            self.append_to_historia(today_str, total_pln)

            return True

        except Exception as e:
            logger_inst.error(f"Failed to update saldo: {e}")
            return False

    def append_to_historia(self, date_str: str, total_pln: float):
        """
        Appends the date and total PLN value to the 'Historia' sheet.
        Creates the sheet if it doesn't exist.
        """
        if self.workbook is None:
            logger_inst.error("Cannot append to Historia: Workbook is not loaded.")
            return

        try:
            if "Historia" not in self.workbook.sheetnames:
                historia_sheet = self.workbook.create_sheet("Historia")
                historia_sheet.cell(row=1, column=1).value = "Date"
                historia_sheet.cell(row=1, column=2).value = "Total_PLN"
                logger_inst.info("Created 'Historia' sheet.")
            else:
                historia_sheet = self.workbook["Historia"]

            # Find the first empty row
            empty_row = 1
            while historia_sheet.cell(row=empty_row, column=1).value is not None:
                empty_row += 1

            historia_sheet.cell(row=empty_row, column=1).value = date_str
            historia_sheet.cell(row=empty_row, column=2).value = total_pln

            logger_inst.info(f"Appended to Historia at row {empty_row}: Date={date_str}, Total_PLN={total_pln:.2f}")

        except Exception as e:
            logger_inst.error(f"Failed to append to Historia: {e}")

    def append_strategy(self, strategy_data: dict) -> bool:
        """
        Appends a new strategy row to the 'Strategia' sheet.
        """
        if self.sheet is None:
            logger_inst.error("Cannot append strategy: Workbook or sheet is not loaded.")
            return False

        try:
            # Find the first empty row below Row 4 (headers on Row 4)
            empty_row = 5
            while self.sheet.cell(row=empty_row, column=1).value is not None:
                empty_row += 1

            today_str = datetime.now().strftime('%Y-%m-%d')

            self.sheet.cell(row=empty_row, column=1).value = strategy_data.get('Wersja', '')
            self.sheet.cell(row=empty_row, column=2).value = today_str
            self.sheet.cell(row=empty_row, column=3).value = strategy_data.get('Klasa aktywów', '')
            self.sheet.cell(row=empty_row, column=4).value = strategy_data.get('Horyzont', '')
            self.sheet.cell(row=empty_row, column=5).value = strategy_data.get('Kryteria wejścia/wyjścia', '')
            self.sheet.cell(row=empty_row, column=6).value = strategy_data.get('Zarządzanie ryzykiem', '')

            logger_inst.info(f"Appended new strategy at row {empty_row}.")
            return True
        except Exception as e:
            logger_inst.error(f"Failed to append strategy: {e}")
            return False

    def append_reflection(self, reflection_data: dict) -> bool:
        """
        Appends a new reflection row to the 'Refleksje' sheet.
        """
        if self.sheet is None:
            logger_inst.error("Cannot append reflection: Workbook or sheet is not loaded.")
            return False

        try:
            # Find the first empty row below Row 4 (headers on Row 4)
            empty_row = 5
            while self.sheet.cell(row=empty_row, column=1).value is not None:
                empty_row += 1

            self.sheet.cell(row=empty_row, column=1).value = "Automatyczne Podsumowanie"
            self.sheet.cell(row=empty_row, column=2).value = reflection_data.get('co_zadzialalo', '')
            self.sheet.cell(row=empty_row, column=3).value = reflection_data.get('co_nie_zadzialalo', '')
            self.sheet.cell(row=empty_row, column=4).value = reflection_data.get('co_zmieniam', '')

            logger_inst.info(f"Appended new reflection at row {empty_row}.")
            return True
        except Exception as e:
            logger_inst.error(f"Failed to append reflection: {e}")
            return False

    @staticmethod
    def get_open_positions_performance(filename: str) -> str:
        """
        Extracts performance data from 'Pozycje otwarte' to format as a string for AI.
        """
        try:
            df = ExcelHandler.get_dashboard_data(filename, "Pozycje otwarte")
            performance = []

            # Look for Instrument (Col C) and Wynik % (Col H in 1-based, might be named different in parsed df)
            # Find columns dynamically based on names
            inst_col = None
            wynik_col = None
            for col in df.columns:
                if 'instrument' in str(col).lower() or 'ticker' in str(col).lower() or 'nazwa' in str(col).lower():
                    inst_col = col
                if 'wynik' in str(col).lower() and '%' in str(col).lower():
                    wynik_col = col

            if not inst_col:
                # Fallback to column index 2 (Column C)
                inst_col = df.columns[2] if len(df.columns) > 2 else None
            if not wynik_col:
                # Fallback to column index 7 (Column H)
                wynik_col = df.columns[7] if len(df.columns) > 7 else None

            if inst_col is None or wynik_col is None:
                logger_inst.warning("Could not find Instrument or Wynik % columns for performance data.")
                return ""

            for _, row in df.iterrows():
                instrument = row.get(inst_col, '')
                wynik = row.get(wynik_col, '')
                if instrument and wynik:
                    performance.append(f"Instrument: {instrument}, Wynik: {wynik}")

            return "\n".join(performance)
        except Exception as e:
            logger_inst.error(f"Failed to extract open positions performance: {e}")
            return ""

    def save_workbook(self) -> bool:
        """
        Saves the workbook to the file.

        Returns:
            bool: True if successful, False otherwise.
        """
        if self.workbook is None:
            logger_inst.error("Cannot save: Workbook is not loaded.")
            return False

        try:
            self.workbook.save(self.filename)
            logger_inst.info(f"Successfully saved changes to {self.filename}")
            return True
        except PermissionError:
            logger_inst.error(
                f"Permission denied while saving {
                    self.filename}. Please close the file if it is open in Excel.")
            return False
        except Exception as e:
            logger_inst.error(f"Unexpected error saving {self.filename}: {e}")
            return False

    def save_sheet_edits(self, data: list) -> bool:
        """
        Saves changes directly to the currently loaded sheet based on the grid data.

        Args:
            data (list): A list of rows, where each row is a list of string values.
                         Assumes the column count matches the data exactly.
        """
        if self.sheet is None:
            logger_inst.error("Cannot save edits: Workbook or sheet is not loaded.")
            return False

        try:
            # Determine start row and skip empty header space based on sheet type
            start_row = 1
            if self.sheet_name in ["Pozycje otwarte", "Trejdy", "Strategia", "Refleksje"]:
                start_row = 5
            elif self.sheet_name == "Salda":
                start_row = 8

            # Clear existing data rows starting from start_row downwards.
            # We determine max row based on actual sheet dimension, and max column.
            max_row = self.sheet.max_row
            max_col = self.sheet.max_column

            for row in range(start_row, max_row + 1):
                for col in range(1, max_col + 1):
                    self.sheet.cell(row=row, column=col).value = None

            # Insert new data
            for r_idx, row_data in enumerate(data):
                # Excel rows are 1-indexed, starting from start_row
                current_row = start_row + r_idx
                for c_idx, val in enumerate(row_data):
                    # Write strings raw/object as per instruction "Treat all incoming edits as raw strings/objects (consistent with our `dtype=object` read strategy)."
                    value_to_write = str(val) if val is not None and val != "" else None
                    self.sheet.cell(row=current_row, column=c_idx + 1).value = value_to_write

            return True

        except Exception as e:
            logger_inst.error(f"Failed to apply sheet edits: {e}")
            return False

    @staticmethod
    def get_dashboard_data(filename: str, sheet_name: str = "Pozycje otwarte") -> pd.DataFrame:
        """
        Reads the specified sheet and returns a clean pandas DataFrame for the dashboard.
        Skips completely empty rows and columns. Handles header rows based on sheet type.
        """
        try:
            skiprows = None
            header = 0

            if sheet_name in ["Pozycje otwarte", "Trejdy", "Strategia", "Refleksje"]:
                skiprows = 3
            elif sheet_name == "Salda":
                skiprows = 6
            elif sheet_name == "Info":
                header = None
            elif sheet_name == "Historia":
                header = 0
                skiprows = None

            df = pd.read_excel(filename, sheet_name=sheet_name, skiprows=skiprows, header=header, dtype=str)

            # Drop completely empty rows and columns
            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # Replace NaNs with empty strings
            df.fillna("", inplace=True)

            return df
        except Exception as e:
            logger_inst.error(f"Error reading dashboard data for sheet '{sheet_name}': {e}")
            return pd.DataFrame()

    @staticmethod
    def export_reports(filename: str) -> bool:
        """
        Exports 'Pozycje otwarte' and 'Trejdy' sheets to CSV files in the exports/ directory.
        """
        export_dir = os.path.join(os.path.dirname(__file__), 'exports')
        os.makedirs(export_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        success = True

        try:
            # Export 'Pozycje otwarte'
            open_pos_sheet = config.get("EXCEL_SHEET_NAME_OPEN_POS", "Pozycje otwarte")
            df_open_pos = pd.read_excel(filename, sheet_name=open_pos_sheet)
            df_open_pos.dropna(how='all', inplace=True)
            open_pos_filename = os.path.join(export_dir, f'portfolio_snapshot_{timestamp}_{open_pos_sheet}.csv')
            df_open_pos.to_csv(open_pos_filename, index=False, encoding='utf-8')
            logger_inst.info(f"Successfully exported {open_pos_sheet} to {open_pos_filename}")
        except Exception as e:
            logger_inst.error(f"Failed to export {open_pos_sheet}: {e}")
            success = False

        try:
            # Export 'Trejdy'
            trades_sheet = config.get("EXCEL_SHEET_NAME_TRADES", "Trejdy")
            df_trades = pd.read_excel(filename, sheet_name=trades_sheet)
            df_trades.dropna(how='all', inplace=True)
            trades_filename = os.path.join(export_dir, f'portfolio_snapshot_{timestamp}_{trades_sheet}.csv')
            df_trades.to_csv(trades_filename, index=False, encoding='utf-8')
            logger_inst.info(f"Successfully exported {trades_sheet} to {trades_filename}")
        except Exception as e:
            logger_inst.error(f"Failed to export {trades_sheet}: {e}")
            success = False

        return success
