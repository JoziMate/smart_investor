import pytest
import tempfile
import os
import openpyxl
from excel_handler import ExcelHandler
from config_manager import config

@pytest.fixture
def temp_excel_file():
    # Create a temporary Excel file
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    # Create workbook with required sheets
    wb = openpyxl.Workbook()

    # Setup Open Positions sheet
    ws_open = wb.active
    ws_open.title = config["EXCEL_SHEET_NAME_OPEN_POS"]
    # Setup trades sheet
    ws_trades = wb.create_sheet(title=config["EXCEL_SHEET_NAME_TRADES"])

    wb.save(path)

    yield path

    # Cleanup
    os.remove(path)

def test_load_workbook_success(temp_excel_file):
    handler = ExcelHandler(temp_excel_file, config["EXCEL_SHEET_NAME_OPEN_POS"])
    assert handler.load_workbook() is True
    assert handler.sheet is not None

def test_load_workbook_failure():
    handler = ExcelHandler("non_existent_file.xlsx", config["EXCEL_SHEET_NAME_OPEN_POS"])
    assert handler.load_workbook() is False

def test_update_asset(temp_excel_file):
    handler = ExcelHandler(temp_excel_file, config["EXCEL_SHEET_NAME_OPEN_POS"])
    handler.load_workbook()

    # Mock ASSET_MAPPING config for testing
    original_mapping = config["ASSET_MAPPING"]
    test_mapping = {'TEST_ASSET': {'row': 2, 'date_col': 'A', 'price_col': 'G'}}

    import excel_handler
    excel_handler.ASSET_MAPPING = test_mapping

    success = handler.update_asset('TEST_ASSET', 123.45)
    assert success is True

    # Verify the values were written
    from openpyxl.utils import column_index_from_string
    from datetime import datetime

    date_col_idx = column_index_from_string('A')
    price_col_idx = column_index_from_string('G')

    assert handler.sheet.cell(row=2, column=date_col_idx).value == datetime.now().strftime('%Y-%m-%d')
    assert handler.sheet.cell(row=2, column=price_col_idx).value == 123.45

    # Restore mapping
    excel_handler.ASSET_MAPPING = original_mapping

def test_append_new_position(temp_excel_file):
    handler = ExcelHandler(temp_excel_file, config["EXCEL_SHEET_NAME_TRADES"])
    handler.load_workbook()

    trade_data = {
        'Platform': 'Test Platform',
        'Ticker': 'TEST',
        'OrderType': 'Market',
        'Direction': 'Long',
        'Volume': 10,
        'Price': 100.5,
        'Justification': 'Test justification'
    }

    success = handler.append_new_position(trade_data)
    assert success is True

    # Verify the values were written to row 5 (first empty row)
    assert handler.sheet.cell(row=5, column=2).value == 'Test Platform'
    assert handler.sheet.cell(row=5, column=3).value == 'TEST'
    assert handler.sheet.cell(row=5, column=7).value == 100.5
