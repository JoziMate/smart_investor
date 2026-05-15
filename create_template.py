import openpyxl

def create_template():
    # Create a new workbook
    wb = openpyxl.Workbook()

    # The default sheet is usually named 'Sheet' or 'Sheet1', rename it to 'Info'
    default_sheet = wb.active
    default_sheet.title = "Info"

    # Create the other required sheets
    wb.create_sheet(title="Strategia")
    wb.create_sheet(title="Salda")
    pozycje_otwarte_sheet = wb.create_sheet(title="Pozycje otwarte")
    trejdy_sheet = wb.create_sheet(title="Trejdy")
    wb.create_sheet(title="Refleksje")

    # Set headers for 'Pozycje otwarte'
    # Column A="Data aktualizacji", Column C="Instrument", Column D="Kierunek",
    # Column E="Wolumen", Column F="Cena wejścia", Column G="Cena bieżąca", Column H="Wynik %"
    pozycje_otwarte_sheet['A1'] = "Data aktualizacji"
    pozycje_otwarte_sheet['C1'] = "Instrument"
    pozycje_otwarte_sheet['D1'] = "Kierunek"
    pozycje_otwarte_sheet['E1'] = "Wolumen"
    pozycje_otwarte_sheet['F1'] = "Cena wejścia"
    pozycje_otwarte_sheet['G1'] = "Cena bieżąca"
    pozycje_otwarte_sheet['H1'] = "Wynik %"

    # Set headers for 'Trejdy'
    # Column A="Data", B="Platforma", C="Instrument", D="Typ zlecenia",
    # E="Kierunek", F="Wolumen", G="Cena", H="Wynik %", I="Uzasadnienie"
    trejdy_headers = ["Data", "Platforma", "Instrument", "Typ zlecenia", "Kierunek", "Wolumen", "Cena", "Wynik %", "Uzasadnienie"]
    for col_idx, header in enumerate(trejdy_headers, start=1):
        trejdy_sheet.cell(row=1, column=col_idx, value=header)

    # Save the workbook to a file
    filename = "Dziennik_inwestora_template.xlsx"
    wb.save(filename)
    print(f"Successfully created template: {filename}")

if __name__ == "__main__":
    create_template()
